from django.shortcuts import render, redirect
from .models import MaintenanceRequest, Machine
from .forms import MaintenanceRequestForm
from dashboard.models import Department
from django.db.models import Q
from .models import Engineer
from .models import SparePart

from django.http import HttpResponse
from openpyxl import Workbook

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.units import inch
from django.http import HttpResponse


def dashboard(request):

    context = {

        # Maintenance Requests
        "total_requests": MaintenanceRequest.objects.count(),
        "open_requests": MaintenanceRequest.objects.filter(status="Open").count(),
        "assigned_requests": MaintenanceRequest.objects.filter(status="Assigned").count(),
        "progress_requests": MaintenanceRequest.objects.filter(status="In Progress").count(),
        "completed_requests": MaintenanceRequest.objects.filter(status="Completed").count(),
        "critical_requests": MaintenanceRequest.objects.filter(priority="Critical").count(),

        # Machines
        "total_machines": Machine.objects.count(),
        "running_machines": Machine.objects.filter(status="Running").count(),
        "breakdown_machines": Machine.objects.filter(status="Breakdown").count(),
        "maintenance_machines": Machine.objects.filter(status="Maintenance").count(),

        # Recent Requests
        "recent_requests": MaintenanceRequest.objects.order_by("-created_at")[:10],
    }

    return render(
        request,
        "maintenance/dashboard.html",
        context,
    )
def request_list(request):

    search = request.GET.get("search", "")
    status = request.GET.get("status", "")
    priority = request.GET.get("priority", "")
    department = request.GET.get("department", "")
    sort = request.GET.get("sort", "newest")

    requests = MaintenanceRequest.objects.all()

    if search:
        requests = requests.filter(
            Q(request_id__icontains=search) |
            Q(machine_name__icontains=search) |
            Q(reported_by__icontains=search)
        )

    if status:
        requests = requests.filter(status=status)

    if priority:
        requests = requests.filter(priority=priority)

    if department:
        requests = requests.filter(department_id=department)

    if sort == "oldest":
        requests = requests.order_by("created_at")
    else:
        requests = requests.order_by("-created_at")

    context = {
        "requests": requests,
        "search": search,
        "status": status,
        "priority": priority,
        "department": department,
        "sort": sort,
        "status_choices": MaintenanceRequest.STATUS_CHOICES,
        "priority_choices": MaintenanceRequest.PRIORITY_CHOICES,
        "departments": Department.objects.all(),
    }

    return render(
        request,
        "maintenance/request_list.html",
        context,
    )
def create_request(request):

    if request.method == "POST":
        form = MaintenanceRequestForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("maintenance_request_list")

    else:
        form = MaintenanceRequestForm()

    return render(
        request,
        "maintenance/request_form.html",
        {
            "form": form,
        },
    )
def request_detail(request, pk):

    maintenance_request = MaintenanceRequest.objects.get(pk=pk)

    return render(
        request,
        "maintenance/request_detail.html",
        {
            "request": maintenance_request,
        },
    )
def edit_request(request, pk):

    maintenance_request = MaintenanceRequest.objects.get(pk=pk)

    if request.method == "POST":

        form = MaintenanceRequestForm(
            request.POST,
            instance=maintenance_request,
        )

        if form.is_valid():
            form.save()

            return redirect(
                "maintenance_request_detail",
                pk=maintenance_request.pk,
            )

    else:

        form = MaintenanceRequestForm(
            instance=maintenance_request,
        )

    return render(
        request,
        "maintenance/request_form.html",
        {
            "form": form,
        },
    )
def delete_request(request, pk):

    maintenance_request = MaintenanceRequest.objects.get(pk=pk)

    if request.method == "POST":
        maintenance_request.delete()

        return redirect(
            "maintenance_request_list"
        )

    return render(
        request,
        "maintenance/request_confirm_delete.html",
        {
            "request": maintenance_request,
        },
    )
from .models import Machine


def machine_list(request):

    search = request.GET.get("search", "")
    status = request.GET.get("status", "")

    machines = Machine.objects.all()

    if search:
        machines = machines.filter(
            Q(machine_code__icontains=search) |
            Q(machine_name__icontains=search) |
            Q(manufacturer__icontains=search)
        )

    if status:
        machines = machines.filter(status=status)

    context = {
        "machines": machines.order_by("machine_code"),
        "search": search,
        "status": status,
        "status_choices": Machine.STATUS_CHOICES,

        "total": Machine.objects.count(),
        "running": Machine.objects.filter(status="Running").count(),
        "breakdown": Machine.objects.filter(status="Breakdown").count(),
        "maintenance": Machine.objects.filter(status="Maintenance").count(),
    }

    return render(
        request,
        "maintenance/machine_list.html",
        context,
    )

def engineer_list(request):

    search = request.GET.get("search", "")
    status = request.GET.get("status", "")

    engineers = Engineer.objects.all()

    if search:
        engineers = engineers.filter(
            Q(engineer_code__icontains=search) |
            Q(engineer_name__icontains=search)
        )

    if status:
        engineers = engineers.filter(status=status)

    context = {

        "engineers": engineers.order_by("engineer_code"),

        "search": search,
        "status": status,

        "status_choices": Engineer.STATUS_CHOICES,

        "total": Engineer.objects.count(),
        "active": Engineer.objects.filter(status="Active").count(),
        "inactive": Engineer.objects.filter(status="Inactive").count(),
    }

    return render(
        request,
        "maintenance/engineer_list.html",
        context,
    )

def sparepart_list(request):

    search = request.GET.get("search", "")
    category = request.GET.get("category", "")

    parts = SparePart.objects.all()

    if search:
        parts = parts.filter(
            Q(part_code__icontains=search) |
            Q(part_name__icontains=search)
        )

    if category:
        parts = parts.filter(category=category)

    low_stock = SparePart.objects.filter(
        stock_quantity__lte=5
    ).count()

    context = {

        "parts": parts.order_by("part_code"),

        "search": search,
        "category": category,

        "category_choices": SparePart.CATEGORY_CHOICES,

        "total": SparePart.objects.count(),

        "low_stock": low_stock,

        "electrical": SparePart.objects.filter(category="Electrical").count(),

        "mechanical": SparePart.objects.filter(category="Mechanical").count(),

    }

    return render(
        request,
        "maintenance/sparepart_list.html",
        context,
    )
def maintenance_report(request):

    status = request.GET.get("status", "")
    priority = request.GET.get("priority", "")
    department = request.GET.get("department", "")

    reports = MaintenanceRequest.objects.all()

    if status:
        reports = reports.filter(status=status)

    if priority:
        reports = reports.filter(priority=priority)

    if department:
        reports = reports.filter(department_id=department)

    context = {
        "reports": reports.order_by("-created_at"),

        "status": status,
        "priority": priority,
        "department": department,

        "status_choices": MaintenanceRequest.STATUS_CHOICES,
        "priority_choices": MaintenanceRequest.PRIORITY_CHOICES,
        "departments": Department.objects.all(),

        "total_reports": reports.count(),
    }

    return render(
        request,
        "maintenance/report.html",
        context,
    )
def export_report_excel(request):

    reports = MaintenanceRequest.objects.all().order_by("-created_at")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Maintenance Report"

    headers = [
        "Request ID",
        "Machine",
        "Department",
        "Status",
        "Priority",
        "Reported By",
        "Assigned Engineer",
        "Created Date",
    ]

    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num).value = header

    row_num = 2

    for report in reports:

        worksheet.cell(row=row_num, column=1).value = report.request_id
        worksheet.cell(row=row_num, column=2).value = report.machine_name
        worksheet.cell(row=row_num, column=3).value = str(report.department)
        worksheet.cell(row=row_num, column=4).value = report.status
        worksheet.cell(row=row_num, column=5).value = report.priority
        worksheet.cell(row=row_num, column=6).value = report.reported_by
        worksheet.cell(row=row_num, column=7).value = report.assigned_engineer
        worksheet.cell(row=row_num, column=8).value = report.created_at.strftime("%d-%m-%Y %H:%M")

        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Maintenance_Report.xlsx"'
    )

    workbook.save(response)

    return response
def export_report_pdf(request):

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="Maintenance_Report.pdf"'
    )

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph("<b>GPMS Maintenance Report</b>", styles["Title"])
    )

    elements.append(
        Paragraph("Generated from GPMS ERP", styles["Normal"])
    )

    elements.append(
        Paragraph("<br/>", styles["Normal"])
    )

    data = [[
        "Request ID",
        "Machine",
        "Department",
        "Status",
        "Priority",
        "Reported By",
    ]]

    reports = MaintenanceRequest.objects.all().order_by("-created_at")

    for report in reports:

        data.append([
            report.request_id,
            report.machine_name,
            str(report.department),
            report.status,
            report.priority,
            report.reported_by,
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,1),(-1,-1),colors.beige),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("BOTTOMPADDING",(0,0),(-1,0),10),

    ]))

    elements.append(table)

    doc.build(elements)

    return response